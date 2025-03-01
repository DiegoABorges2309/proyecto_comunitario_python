from datetime import datetime
import time
from openpyxl import load_workbook

class Functions:
    def __init__(self):
        pass

    def group_list(self, c_list):
        for index in range(5):
            if c_list[index] == '':
                c_list[index] = None
        if c_list[1] != None:
            c_list[1] = float(c_list[1])

        if c_list[4] != None:
            c_list[4] = datetime.strptime(c_list[4], "%Y-%m-%d").date()

        return c_list

class SaveExel:
    def __init__(self):
        self.result_boolean = None
        self._index = 0

    def save_one(self, _row, _quantity, _wb, _file):
        self._index = 2
        for _cell in _row[-self._index::]:
            _cell.value = _quantity
            _wb.save(_file)
            return True

    def save_two(self, _row, _quantity, _time, _wb, _file):
        self._index = 21
        for _cell in _row[-self._index::]:
            _cell.value = _quantity
            _wb.save(_file)
            break

        for _cell_two in _row[-self._index+1::]:
            _cell_two.value = _time
            _wb.save(_file)
            break

    def save_tre(self, _row, _quantity, _wb, _file):
        self._index = 3
        for _cell in _row[-self._index::]:
            _cell.value = _quantity
            _wb.save(_file)
            return True

    def save(self, _file, _sheet, _column, _name, quantity):
        wb = load_workbook(_file)
        ws = wb[_sheet]
        column = ws[_column]
        for cell in column:
            if cell.value is not None and cell.value == _name:
                _row = ws[cell.row]
                if _file == "docx/docxxx1.xlsx":
                    SaveExel.save_one(self, _row, quantity, wb, _file)

                if _file == "docx/docxxx2.xlsx":
                    time_today = datetime.today()
                    convert_time = datetime.strptime(str(time_today.date()), "%Y-%m-%d").date()
                    SaveExel.save_two(self, _row, quantity, convert_time.strftime(format="%d/%m/%Y"), wb, _file)

                if _file == "docx/docxxx3.xlsx":
                    SaveExel.save_tre(self, _row, quantity, wb, _file)

                return True
            else:
                pass

    def action_save(self, _name, _quantity):
        list_name = _name
        list_file = ["docx/docxxx1.xlsx", "docx/docxxx2.xlsx", "docx/docxxx3.xlsx"]
        list_sheet = ['MMQ  HOSPITAL ', 'MATERIA MQ', 'CONSOLIDADO MMQ']
        list_column = ['A', 'B', 'B']
        clean_sheet = 'MATERIAL DE ASEO-LIMPIEZA '
        self.result_boolean = False
        for index in range(3):
            _bool = self.save(list_file[index], list_sheet[index], list_column[index], list_name[index], _quantity)
            if _bool is not None:
                self.result_boolean = _bool

        return self.result_boolean
